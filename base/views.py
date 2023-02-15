from django.shortcuts import render
from .models import (Gallery,Team,logo,Carrer,blog,Testimonials,Events,HowWeWork,Birac,Tbi,DemoDayTOPSECTION,DemoDayPic,CentralGovernmentFundingDB,
                    Start_UpTN,StateGovtFund,OurStartup,Investors,AboutHeading,MBADB,Govt_Tie,LastContent,UploadImage,GlobalMarket,GlobalMarketPic,stateGovtFunddb,
                    Start_UpTNContent2,StateGovtFundSecondSection,International_Partners,Sisfs,WhoAreWe,Contact_SECTION,HOME_TESTIMONIAL,EventsForm,Facilities_developed,About_SISFS,
                    Start_UpTNimg1,Start_UpTNimg2,StateGovtFundEligibilitySection,MentorConnectDB,MentorClinicDB, angelInvestorDB, new_venturesDB,TOPSECTION,WhatWeDo,OurProcess,SpendingSection,JoinOurCommunity)
from .Tools import get_images,get_team,reguler_datas,get_blog,get_startup,get_DemoDayPic
import datetime
import json
import openpyxl

# Create your views here.


def admin(request):
    item = get_images()
    return render(request,"pages/empty.html",reguler_datas({"categories":item[0],"images":item[1]}))


#...............gallery.......................................
def gallery(request):
    item = get_images()
    return render(request,"gallery.html",reguler_datas({"categories":item[0],"images":item[1]}))
#............................................................
# upload image...............................................
def upload_image(request):
    categories = request.POST.get("Category")
    image = request.FILES["image_file"]
    update = Gallery(image=image,categories=categories)
    update.save()
    return render(request,"about_us/team.html")

def delete_image(request):
    id = request.POST.get("id")
    image = Gallery.objects.get(G_id=id)
    image.delete()
    return render(request,"about_us/team.html")
#..............................................................

def aboutus(request):
    return render(request,"about_us/team.html",reguler_datas({"team":get_team()}))


# upload team ...............................
def admin_team(request):
    return render(request,'pages/team.html',reguler_datas({"team":get_team()}))

def update_team(request):
    name = request.POST.get("name")
    Category = request.POST.get("Category")
    position = request.POST.get("position")
    image = request.FILES["image_file"]
    db_obj = Team(Name=name,categories=Category,image=image,position=position)
    db_obj.save()
    obj = Team.objects.all()
    for i in obj:
        print(i.Name,i.image)

    return render(request,'pages/team.html')

def delete_team(request):
    id = request.POST.get("id")
    image = Team.objects.get(Team_id=id)
    image.delete()
    return render(request,"about_us/team.html",reguler_datas())

#............................................................
#...............Logo.........................................
def update_logo(request):
    return render(request,"home/logo.html",reguler_datas())

def upload_logo(request):
    Reson = request.POST.get("#reson")
    image = request.FILES["#fileInput-single"]
    update = logo(image=image,Reson=Reson)
    update.save()
    return render(request,"home/logo.html")

def delete_logo(request):
    id = request.POST.get("id")
    size = len(logo.objects.all())
    if size > 1:
        image = logo.objects.get(L_id=id)
        image.delete()
    else:
        pass
    return render(request,"home/logo.html")

def set_logo(request):
    id = request.POST.get("id")
    image = logo.objects.get(L_id=id)
    image_ = image.image
    reson = image.Reson
    image.delete()
    obj = logo(image=image_,Reson=reson)
    obj.save()
    return render(request,"home/logo.html")
#............................................................
#...............Logo.........................................

def carrer(request):
    return render(request,"about_us/carrer.html")
def update_carrer(request):
    ids=['#name','#Qualification','#Experience','#Email','#Subject','#Message']
    Name = request.POST.get(ids[0])
    Email = request.POST.get(ids[3])
    Message = request.POST.get(ids[-1])
    Subject = request.POST.get(ids[4])
    qualififcation = request.POST.get(ids[1])
    experience = request.POST.get(ids[2])
    image = request.FILES["#fileInput-single"]
    obj = Carrer(Name=Name,Email=Email,Message=Message,Subject=Subject,qualififcation=qualififcation,experience=experience,image=image)
    obj.save()
    ob=Carrer.objects.all()
    for i in ob:
        print(i.Name)
    return render(request,"about_us/carrer.html")
    
#............................................................

#...............Blog........................................
def blog_edit(request):
    return render(request,"pages/blog_edit.html")

def save_blog(request):
    ids = ['#title','#description','#content','#Category','#Thumbnail']
    title = request.POST.get(ids[0])
    description = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    Category = request.POST.get(ids[3])
    Thumbnail = request.POST.get(ids[4])

    obj = blog(title=title,description=description,content=content,categories=Category,blog_profile_img=Thumbnail)
    obj.save()
    ob = blog.objects.all()
    for i in ob:
        print(i.blog_profile_img,i.title,i.content)

    return render(request,"pages/blog_edit.html")

def save_edit_blog(request,pk):
    ids = ['#title','#description','#content','#Category','#Thumbnail']
    title = request.POST.get(ids[0])
    description = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    Category = request.POST.get(ids[3])
    Thumbnail = request.POST.get(ids[4])

    obj = blog.objects.get(id=pk)
    obj.content = content
    obj.title = title
    obj.description = description
    obj.categories = Category
    obj.blog_profile_img = Thumbnail
    obj.save()

    print("Saved...........")

    return render(request,"pages/blog_edit.html")


def list_blog(request):
    items = get_blog()
    return render(request,"home/blog.html",{'blogs':items})

def view_blog(request,pk):
    page = blog.objects.get(id=pk)
    items = get_blog()
    return render(request,"home/view_blog.html",{'blog':page,'item':items})

def delete_blog(request):
    bl_id = request.GET.get("id")
    page = blog.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})


def list_edit_blog(request):
    items = get_blog()
    return render(request,"home/edit_blog_list.html",{'blogs':items})

def edit_blog(request,pk):
    obj = blog.objects.get(id=pk)
    return render(request,"pages/blog_re_edit.html",{'obj':obj})


#............................................................

#.....................Events.........................

def events(request):
    obj = Events.objects.all()
    return render(request,"about_us/Events.html",reguler_datas({"card":obj}))

def events_edit(request):
    obj = Events.objects.all()
    return render(request,"pages/Events_edit.html",reguler_datas({"card":obj}))


def events_save(request):
    ids = ['#topic','#program','#date','#Category','#fileInput-single']
    description = request.POST.get(ids[1])
    topic = request.POST.get(ids[0])
    date = request.POST.get(ids[2])
    categories = request.POST.get(ids[3])
    image = request.FILES[ids[-1]]
    d = date.split("-")
    date_formate = datetime.date(int(d[0]), int(d[1]), int(d[2]))
    obj = Events(description=description,type=topic,categories=categories,date=date_formate,image=image)
    obj.save()
    print("....................saved.......")
    return render(request,"about_us/Events.html",reguler_datas({"card":obj}))


#............................................................


#.....................Testimonicals.........................

def Testimonicals(request):
    obj = Testimonials.objects.all()
    return render(request,"about_us/Testimonicals.html",reguler_datas({"card":obj}))

def Testimonicals_edit(request):
    obj = Testimonials.objects.all()
    return render(request,"pages/Testimonicals_edit.html",reguler_datas({"card":obj}))

def Testimonicals_save(request):
    vals = ['#Name','#position','#description','#Category','#fileInput-single']
    Name = request.POST.get(vals[0])
    position = request.POST.get(vals[1])
    image = request.FILES[vals[-1]]
    description = request.POST.get(vals[2])
    categories = request.POST.get(vals[3])

    vals = Testimonials(Name=Name,position=position,image=image,description=description,categories=categories)
    vals.save()

    return render(request,"pages/Testimonicals_edit.html")

#............................................................

#...............birac........................................
def birac(request):
    obj = Birac.objects.latest('id')
    re_view = obj.overview.split(',')
    topic= {}
    for i in re_view:
        if '~' in i:
           topic[i.split('~')[1]] = i.split('~')[0]
            
    print(topic)

    return render(request,"about_us/birac.html",reguler_datas({'birac':obj,"topic":topic}))

def birac_edit(request):
    obj = Birac.objects.all()[::-1]
    item = []
    topic= {}
    birac_ = {}
    for i in obj:
        topic=dict()
        for j in i.overview.split(','):
             if '~' in j:
                topic[j.split('~')[1]] = j.split('~')[0]
        item.append(topic)
    print(item)
    for i,x in enumerate(obj):
        birac_[x] = item[i]
    print(birac_)
    return render(request,"pages/birac_edit.html",reguler_datas({'birac':birac_,'view':item,'Facilities_developed':Facilities_developed.objects.all()}))

def birac_save(request):
    ids = ['#title','#subtitle','#content','tags']
    title =request.POST.get(ids[0])
    subtitle_ = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    tags = request.POST.get(ids[3])
    item=""
    print(subtitle_)
    res = json.loads(tags)
    for i in res:
        item = item + i.get('value') + ", "
    try:
        image = request.FILES['#fileInput-single_image']
        obj = Birac(title=title,subtitle=subtitle_,description=content,overview=item,image=image)
        obj.save()
    except:
        obj = Birac.objects.all()[::-1][0]
        obj = Birac(title=title,subtitle=subtitle_,description=content,overview=item,image=obj.image)
        obj.save()
    print("/..........................................saved................................................../")
    return render(request,"pages/birac_edit.html",reguler_datas())

def delete_birac(request):
    bl_id = request.GET.get("id")
    page = Birac.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def set_birac(request):
    id_ = request.GET.get("id")
    image = Birac.objects.get(id=id_)
    title = image.title
    subtitle = image.subtitle
    description = image.description
    overview = image.overview
    images = image.image
    image.delete()
    obj = Birac(title=title,subtitle=subtitle,description=description,overview=overview,image=images)
    obj.save()
    print("saved")
    return render(request,"home/logo.html")


def facilities_developed_save(request):
    ids = ['#Heading','#Content']
    Heading =request.POST.get(ids[0])
    Content = request.POST.get(ids[1])
    
    obj = Facilities_developed(topic=Heading,content=Content)
    obj.save()
    return render(request,"home/logo.html")

def delete_facilities_developed(request):
    bl_id = request.GET.get("id")
    page = Facilities_developed.objects.get(FD_id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def SISFS_scheme_save(request):
    ids = ['#Content']
    Content = request.POST.get(ids[0])
    obj = About_SISFS(topic=Content)
    obj.save()
    return render(request,"home/logo.html")

def delete_SISFS_scheme(request):
    bl_id = request.GET.get("id")
    page = About_SISFS.objects.get(FD_id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

#............................................................

#...............tbi........................................
def tbi(request):
    try:
        obj = Tbi.objects.latest('id')
        re_view = obj.overview.split(',')
        topic= {}
        for i in re_view:
            if '~' in i:
                topic[i.split('~')[1]] = i.split('~')[0]
    except:
        obj = ""
        topic = ""
        print("maybe the database are empty.....")
            
    print(topic)

    return render(request,"about_us/tbi.html",reguler_datas({'birac':obj,"topic":topic}))

def tbi_edit(request):
    obj = Tbi.objects.all()
    item = []
    topic= {}
    tbi_ = {}
    for i in obj:
        topic=dict()
        for j in i.overview.split(','):
             if '~' in j:
                topic[j.split('~')[1]] = j.split('~')[0]
        item.append(topic)
    print(item)
    for i,x in enumerate(obj):
        tbi_[x] = item[i]
    print(tbi_)
    return render(request,"pages/tbi_edit.html",reguler_datas({'birac':tbi_,'view':item}))

def tbi_save(request):
    ids = ['#title','#subtitle','#content','tags']
    title =request.POST.get(ids[0])
    subtitle_ = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    tags = request.POST.get(ids[3])
    item=""
    print(subtitle_)
    res = json.loads(tags)
    for i in res:
        item = item + i.get('value') + ", "
    obj = Tbi(title=title,subtitle=subtitle_,description=content,overview=item)
    obj.save()
    print("/..........................................saved................................................../")
    return render(request,"pages/tbi_edit.html",reguler_datas())

def delete_tbi(request):
    bl_id = request.GET.get("id")
    page = Tbi.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def set_tbi(request):
    id_ = request.GET.get("id")
    image = Tbi.objects.get(id=id_)
    title = image.title
    subtitle = image.subtitle
    description = image.description
    overview = image.overview
    image.delete()
    obj = Tbi(title=title,subtitle=subtitle,description=description,overview=overview)
    obj.save()
    print("saved")
    return render(request,"home/logo.html")

#............................................................

#...............sisfs........................................
def sisfs(request):
    try:
        obj = Sisfs.objects.latest('id')
        re_view = obj.overview.split(',')[0:-1]
        print(re_view)
    except:
        obj = ""
        re_view = ""
        print("maybe the database are empty.....")

    return render(request,"sisfs.html",reguler_datas({'birac':obj,"topic":re_view,'data':Sisfs.objects.latest('id'),'abt':About_SISFS.objects.all()}))

def sisfs_edit(request):
    obj = Sisfs.objects.all()[::-1]
    item = []
    topic = list()
    sisfs_ = {}
    for i in obj:
        topic=list()
        for j in i.overview.split(','):
                topic.append(j)
        item.append(topic)
    print(item)
    for i,x in enumerate(obj):
        sisfs_[x] = item[i]
    print(sisfs_)
    return render(request,"pages/sisfs_edit.html",reguler_datas({'birac':sisfs_,'view':item,'sisfs':Sisfs.objects.all(),'About_SISFS':About_SISFS.objects.all()}))

def sisfs_save(request):
    ids = ['#title','#subtitle','#content','tags']
    title =request.POST.get(ids[0])
    subtitle_ = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    tags = request.POST.get(ids[3])
    item=""
    print(subtitle_)
    res = json.loads(tags)
    for i in res:
        item = item + i.get('value') + ", "
    try:
        image = request.FILES['#fileInput-single_image']
        obj = Sisfs(title=title,subtitle=subtitle_,description=content,overview=item,image=image)
        obj.save()
    except:
        obj = Sisfs.objects.all()[::-1][0]
        obj = Sisfs(title=title,subtitle=subtitle_,description=content,overview=item,image=obj.image)
        obj.save()

    print("/..........................................saved................................................../")
    return render(request,"pages/sisfs_edit.html",reguler_datas())

def delete_sisfs(request):
    bl_id = request.GET.get("id")
    page = Sisfs.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def set_sisfs(request):
    id_ = request.GET.get("id")
    image = Sisfs.objects.get(id=id_)
    title = image.title
    subtitle = image.subtitle
    description = image.description
    overview = image.overview
    images = image.image
    image.delete()
    obj = Sisfs(title=title,subtitle=subtitle,description=description,overview=overview,image=images)
    obj.save()
    print("saved")
    return render(request,"home/logo.html")





def eventform(request):
    return render(request,"about_us/event_form.html")

def edit_eventform(request):
    obj = EventsForm.objects.all()
    return render(request,"pages/eventform_edit.html",{'obj':obj})

def update_eventform(request):
    title = request.POST.get('#title')
    Name = request.POST.get('#name')
    Email = request.POST.get('#Email')
    company = request.POST.get('#company')
    event = request.POST.get('#event')
    linkedin = request.POST.get('#linkedin')
    website  =request.POST.get('#website')
    image = request.FILES["#fileInput-single"]
    obj = EventsForm(title=title,Name=Name,Email=Email,linkedin=linkedin,website=website,company=company,event=event,image=image)
    obj.save()
    ob=EventsForm.objects.all()
    for i in ob:
        print(i.Name)
    return render(request,"about_us/carrer.html")

def delete_form(request):
    bl_id = request.GET.get("id")
    page = EventsForm.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def convert_excel(request):
    wb = openpyxl.Workbook() 
    sheet = wb.active 
    data = EventsForm.objects.all()
    title = ["updated_date","title","Name","Email","company","event","linkedin","website"]
    for i,x in enumerate(title):
        cell_obj = sheet.cell(row = 1, column = i+1)
        cell_obj.value = x

    for i,x in enumerate(data,2):
        row_data = [x.updated_date,x.title,x.Name,x.Email,x.company,x.event,x.linkedin,x.website]
        for j,y in enumerate(row_data):
            cell_obj = sheet.cell(row = i+1, column = j+1)
            cell_obj.value = y
            print(i,j)
    wb.save("sample5.xlsx") 
    return render(request,"home/view_blog.html")


def EDI (request):
    return render(request,"fund/edi.html")


def angelInvestor (request):
    return render(request,"angelinvestor.html")

def home(request):
    return render(request,"index.html")

def MentorClinic (request):
    return render(request,"mentorclinic.html")



def MentorConnect (request):
    return render(request,"mentorconnect.html",{'mentor':MentorConnectDB.objects.all()[::-1]})

def MontorConnect_edit(request):
    return render(request,"edtior/mentor_connect_edit.html",{'mentor':MentorConnectDB.objects.all()[::-1]})

def MontorConnect_save(request):
    content = request.POST.get('#content')
    if request.POST.get('#page') == 'MentorConnect':
        obj = MentorConnectDB(Content=content)
        obj.save()
    print("saved.....................................////////////////////////")
    return render(request,"edtior/mentor_connect_edit.html")



def MentorClinic (request):
    return render(request,"mentorclinic.html")

def Mentor_Clinic_edit(request):
    return render(request,"edtior/Mentor_Clinic_edit.html",{'mentor':MentorClinicDB.objects.all()[::-1]})

def Mentor_Clinic_save(request):
    content = request.POST.get('#content')
    obj = MentorClinicDB(Content=content)
    obj.save()
    print("saved.....................................////////////////////////")
    return render(request,"edtior/Mentor_Clinic_edit.html")


def angelInvestor (request):
    return render(request,"angelInvestor.html",{'mentor':angelInvestorDB.objects.all()[::-1]})

def angelInvestor_edit(request):
    return render(request,"edtior/angelInvestor_edit.html",{'mentor':angelInvestorDB.objects.all()[::-1]})

def angelInvestor_save(request):
    content = request.POST.get('#content')
    obj = angelInvestorDB(Content=content)
    obj.save()
    for i in angelInvestorDB.objects.all():
        print(i.Content)
    print("saved.....................................////////////////////////")
    return render(request,"edtior/angelInvestor_edit.html")

def new_ventures (request):
    data = new_venturesDB.objects.all()[::-1]
    print(data)
    return render(request,"newventures.html",{'mentor':data,'sample':'hi'})

def new_ventures_edit(request):
    return render(request,"edtior/new_ventures_edit.html",{'mentor':new_venturesDB.objects.all()[::-1]})

def new_ventures_save(request):
    content = request.POST.get('#content')
    obj = new_venturesDB(Content=content)
    obj.save()
    for i in new_venturesDB.objects.all():
        print(i.Content)
    print("saved.....................................////////////////////////")
    return render(request,"edtior/new_ventures_edit.html")


def samridth (request):
    return render(request,"samridth.html")

def testimonial (request):
    return render(request,"testimonials.html")

def career (request):
    return render(request,"carrer.html")

def home(request):
    try :
        whoweare = WhoAreWe.objects.all()[::-1]
        home_TESTIMONIAL =  HOME_TESTIMONIAL.objects.all()[::-1]
        contact_Section =  Contact_SECTION.objects.all()[::-1]
        investors = Investors.objects.all()[::-1]
        govt = Govt_Tie.objects.all()[::-1]
        Uploadimage = UploadImage.objects.all()[::-1]

        
        Internationalpartners = International_Partners.objects.all()[::-1]
        return render(request,"index.html",{'whoweare':whoweare,'ht':home_TESTIMONIAL,'cs':contact_Section,'investors':investors,'ip':Internationalpartners,'govt':govt,'Uploadimage':Uploadimage})
    except:
        print("maybe database are empty")
    return render(request,"pages/home_edit.html")


def home_edit(request):
    try :
        whoweare = WhoAreWe.objects.all()[::-1]
        home_TESTIMONIAL =  HOME_TESTIMONIAL.objects.all()[::-1]
        contact_Section =  Contact_SECTION.objects.all()[::-1]
        investors = Investors.objects.all()[::-1]
        Internationalpartners = International_Partners.objects.all()[::-1]
        govt = Govt_Tie.objects.all()[::-1]
        Uploadimage = UploadImage.objects.all()[::-1]


        return render(request,"pages/home_edit.html",{'whoweare':whoweare,'ht':home_TESTIMONIAL,'cs':contact_Section,'investors':investors,'ip':Internationalpartners,'govt':govt,'Uploadimage':Uploadimage})
    except:
        print("maybe database are empty")
    return render(request,"pages/home_edit.html")


def Whoweare(request):
    ids = ['#Sub-heading','#Point-1','#Point-2','#Point-3','#Point-4','#image_file']
    SubHeading = request.POST.get(ids[0])
    Point1 = request.POST.get(ids[1])
    Point2 = request.POST.get(ids[2])
    Point3 = request.POST.get(ids[3])
    Point4 = request.POST.get(ids[4])

    try :
        image = request.FILES[ids[5]]
    except:
        image = WhoAreWe.objects.all()[::-1][0].image

    obj = WhoAreWe(SubHeading=SubHeading,Point1=Point1,Point2=Point2,Point3=Point3,Point4=Point4,image=image)
    obj.save()

    for i in WhoAreWe.objects.all():
        print(i.SubHeading)

    return render(request,"gallery.html")


def  Home_TESTIMONIAL(request):
    ids = ['#fileInput-single1','#Testimonial','#Name','#Designation']

    Testimonial_content = request.POST.get(ids[1])
    Name = request.POST.get(ids[2])
    Designation = request.POST.get(ids[3])
    try :
        image = request.FILES[ids[0]]
    except:
        image = HOME_TESTIMONIAL.objects.all()[::-1][0].image

    obj = HOME_TESTIMONIAL(Testimonial_content=Testimonial_content,Name=Name,Designation=Designation,image=image)
    obj.save()

    for i in HOME_TESTIMONIAL.objects.all():
        print(i.Name)

    return render(request,"gallery.html")

def  Contact_Section(request):
    ids = ['#Title','#Address','#pno','#E-Mail']

    Title = request.POST.get(ids[0])
    Address = request.POST.get(ids[1])
    Phone_number = request.POST.get(ids[2])
    E_Mail = request.POST.get(ids[3])

    obj = Contact_SECTION(Title=Title,Address=Address,Phone_number=Phone_number,E_Mail=E_Mail)
    obj.save()

    for i in Contact_SECTION.objects.all():
        print(i.Title)

    return render(request,"gallery.html")

def investors(request):
    ids = ['#fileInput-single2','#sub-heading']

    heading = request.POST.get(ids[1])
    try :
        image = request.FILES[ids[0]]
        obj = Investors(Title=heading,image=image)
        obj.save()
    except:
        image = Investors.objects.all()[::-1][0]
        image.Title = heading
        image.save()


    for i in Investors.objects.all():
        print(i.Title)

    return render(request,"gallery.html")

def International(request):
    ids = ['#fileInput-single3','#sub-heading6']

    heading = request.POST.get(ids[1])
    try :
        image = request.FILES[ids[0]]
        obj = International_Partners(Title=heading,image=image)
        obj.save()
    except:
        image = International_Partners.objects.all()[::-1][0]
        image.Title = heading
        image.save()


    for i in International_Partners.objects.all():
        print(i.Title)

    return render(request,"gallery.html")

def GovtTie(request):
    ids = ['#fileInput-single4','61']

    heading = request.POST.get(ids[1])
    try :
        image = request.FILES[ids[0]]
        obj = Govt_Tie(Title=heading,image=image)
        obj.save()
    except:
        image = Govt_Tie.objects.all()[::-1][0]
        image.Title = heading
        image.save()

    for i in Govt_Tie.objects.all():
        print(i.Title)

    return render(request,"gallery.html")

def Upload_Image(request):
    ids = ['#fileInput-single8','#Appear','#Appear-image']

    heading = request.POST.get(ids[1])
    content = request.POST.get(ids[2])

    try :
        image = request.FILES[ids[0]]
        obj = UploadImage(Title=heading,image=image,content=content)
        obj.save()
    except:
        image = UploadImage.objects.all()[::-1][0]
        image.Title = heading
        image.save()

    for i in UploadImage.objects.all():
        print(i.Title)

    return render(request,"gallery.html")



def MBA (request):
    return render(request,"mba.html",{'mentor':MBADB.objects.all()[::-1]})

def MBA_edit(request):
    return render(request,"edtior/MBA_edit.html",{'mentor':MBADB.objects.all()[::-1]})

def MBA_save(request):
    content = request.POST.get('#content')
    obj = MBADB(Content=content)
    obj.save()
    print("saved.....................................////////////////////////")
    return render(request,"edtior/MBA_edit.html")

def ourStartups (request):
    return render(request,"ourstartup.html",{'startup':get_startup()})

def ourStartups_edit (request):
    return render(request,"pages/ourstartup_edit.html",{'startup':get_startup()})

def ourStartups_save (request):
    Sub_heading = request.POST.get('#Sub_heading')
    Name = request.POST.get('#Name')
    content = request.POST.get('#Content')
    categories = request.POST.get('#Categories_for_startup')
    image = request.FILES['#fileInput-single']
    obj = OurStartup(Sub_heading=Sub_heading,Name=Name,content=content,categories=categories,image=image)
    obj.save()
    for i in OurStartup.objects.all():
        print(i.Name)

    return render(request,"pages/ourstartup_edit.html")
def delete_startup(request):
    bl_id = request.POST.get("id")
    page = OurStartup.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})


def about (request):
    return render(request,"a/thiruabout.html",{'about_heading':AboutHeading.objects.all()[::-1],'HowWeWork':HowWeWork.objects.all()[::-1],'last_con':LastContent.objects.all()[::-1]})

def about_edit (request):
    return render(request,"pages/about_edit.html",{'about_heading':AboutHeading.objects.all()[::-1],'HowWeWork':HowWeWork.objects.all()[::-1],'last_con':LastContent.objects.all()[::-1]})


def HowWeWork_save (request):
    Point_1 = request.POST.get("#Point-1")
    Sub_heading = request.POST.get("#Sub-heading")
    Point_2 = request.POST.get("#Point-2")
    Point_3 = request.POST.get("#Point-3")
    Point_4 = request.POST.get("#Point-4")

    try :
        image = request.FILES['#fileInput-single']
        obj = HowWeWork(Sub_heading=Sub_heading,Point_1=Point_1,Point_2=Point_2,Point_4=Point_4,Point_3=Point_3,image=image)
        obj.save()
    except:
        obj = HowWeWork.objects.all()[::-1][0]
        obj = HowWeWork(Sub_heading=Sub_heading,Point_1=Point_1,Point_2=Point_2,Point_4=Point_4,Point_3=Point_3,image=obj.image)
        obj.save()

    return render(request,"pages/about_edit.html",{})


def Last_content_save (request):
    Sub_heading = request.POST.get("#sub_heading")
    para_1 = request.POST.get("#para_1")
    para_2 = request.POST.get("#para_2")
    heading_1 = request.POST.get("#heading_1")
    sub_para_1 = request.POST.get("#sub_para_1")
    heading_2 = request.POST.get("#heading_2")
    sub_para_2 = request.POST.get("#sub_para_2")
    heading_3 = request.POST.get("#heading_3")
    sub_para_3 = request.POST.get("#sub_para_3")

    try:
        image = request.FILES['#image_file_upload']
        obj = LastContent(image=image,Sub_heading=Sub_heading,sub_para_3=sub_para_3,heading_3=heading_3,heading_2=heading_2,sub_para_2=sub_para_2,para_1=para_1,para_2=para_2,heading_1=heading_1,sub_para_1=sub_para_1)
        obj.save()
    except:
        obj = LastContent.objects.all()[::-1][0]
        obj = LastContent(image=obj.image,Sub_heading=Sub_heading,sub_para_3=sub_para_3,heading_3=heading_3,heading_2=heading_2,sub_para_2=sub_para_2,para_1=para_1,para_2=para_2,heading_1=heading_1,sub_para_1=sub_para_1)
        obj.save()

    
    return render(request,"pages/about_edit.html",{})

def About_heading_save (request):
    ['#SUb_heading','#Detail-text']
    Sub_heading = request.POST.get("#SUb_heading")
    Detail_text = request.POST.get("#Detail-text")

    obj=AboutHeading(Sub_heading=Sub_heading,Detail_text=Detail_text)
    obj.save()

    return render(request,"pages/about_edit.html",{})


def Global_Market (request):
    return render(request,"globalmarket.html",{'pics':GlobalMarketPic.objects.all()[::-1],'title':GlobalMarket.objects.all()[::-1]})

def GlobalMarket_edit (request):
    return render(request,"pages/global_market_edit.html",{'pics':GlobalMarketPic.objects.all()[::-1],'title':GlobalMarket.objects.all()[::-1]})

def GlobalMarket_save (request):
    Heading = request.POST.get("#heading")
    Content = request.POST.get("#Content")
    obj = GlobalMarket(Heading=Heading,Content=Content)
    obj.save()
    return render(request,"pages/global_market_edit.html")

def GlobalMarketPic_save (request):
    title = request.POST.get("#pic_title")
    image = request.FILES["#image_load"]
    print(title)
    obj = GlobalMarketPic(title=title,image=image)
    obj.save()
    print("................................saved...........................................")
    return render(request,"pages/global_market_edit.html")

def delete_GlobalMarketPic(request):
    id = request.POST.get("id")
    size = len(GlobalMarketPic.objects.all())
    if size > 1:
        image = GlobalMarketPic.objects.get(id=id)
        image.delete()
    else:
        pass
    return render(request,"home/logo.html")

def set_GlobalMarketPic(request):
    id = request.POST.get("id")
    image = GlobalMarketPic.objects.get(id=id)
    image_ = image.image
    reson = image.title
    image.delete()
    obj = GlobalMarketPic(image=image_,title=reson)
    obj.save()
    return render(request,"home/logo.html")

def service (request):
    return render(request,"service.html",{'JoinOurCommunity':JoinOurCommunity.objects.all()[::-1],'SpendingSection':SpendingSection.objects.all()[::-1],'OurProcess':OurProcess.objects.all()[::-1],'top':TOPSECTION.objects.all()[::-1],'WhatWeDo':WhatWeDo.objects.all()[::-1]})

def service_edit (request):
    return render(request,"pages/service_edit.html",{'JoinOurCommunity':JoinOurCommunity.objects.all()[::-1],'SpendingSection':SpendingSection.objects.all()[::-1],'OurProcess':OurProcess.objects.all()[::-1],'top':TOPSECTION.objects.all()[::-1],'WhatWeDo':WhatWeDo.objects.all()[::-1]})

def TopSection_save(request):
    Sub_Heading = request.POST.get("#sub_heading")
    Heading = request.POST.get("#heading")

    try:
        image = request.FILES['#image_upload1']
        obj = TOPSECTION(Heading=Heading,Sub_Heading=Sub_Heading,image=image)
        obj.save()
    except:
        obj = TOPSECTION.objects.all()[::-1][0]
        obj = TOPSECTION(Heading=Heading,Sub_Heading=Sub_Heading,image=obj.image)
        obj.save()

    obj.save()
    print("saved")
    return render(request,"pages/service_edit.html")

def WhatWedo_save(request):
    Sub_Heading = request.POST.get("#Sub-heading")
    Para_below_heading = request.POST.get("#Point-1")
    para_27 = request.POST.get("#Point-2")
    Secure_Payment_para = request.POST.get("#Point-3")
    Daily_Update_para = request.POST.get("#Point-4")
    Market_Research_para = request.POST.get("#Point-5")

    obj = WhatWeDo(Sub_Heading=Sub_Heading,Para_below_heading=Para_below_heading,para_27=para_27,Secure_Payment_para=Secure_Payment_para,Daily_Update_para=Daily_Update_para,Market_Research_para=Market_Research_para)
    obj.save()
    return render(request,"pages/service_edit.html")

['#sub_heading','#para_1','#para_2','#heading_1']

def Our_Process_save(request):
    heading = request.POST.get("#sub_heading")
    concept_para = request.POST.get("#para_1")
    prepare_para = request.POST.get("#para_2")
    retouch_para = request.POST.get("#heading_1")
    obj = OurProcess(heading=heading,concept_para=concept_para,prepare_para=prepare_para,retouch_para=retouch_para)
    obj.save()

    return render(request,"pages/service_edit.html")

def Spending_Section_save(request):
    heading = request.POST.get("#Sub-heading")
    point3 = request.POST.get("#para_1")
    Para_below_heading = request.POST.get("#Para_below_heading")
    Point_1 = request.POST.get("#Point-1")
    Point_2 = request.POST.get("#Point-2")
    Point_3 = request.POST.get("#Point-3")
    Point_4 = request.POST.get("#Point-4")
    
    try:
        image = request.FILES['#upload_image1']
        obj = SpendingSection(image=image,heading=heading,Para_below_heading=Para_below_heading,Point_1=Point_1,Point_2=Point_2,Point_3=Point_3,Point_4=Point_4)
        obj.save()
    except:
        obj = SpendingSection.objects.all()[::-1][0]
        obj = SpendingSection(image=obj.image,heading=heading,point3=point3,Para_below_heading=Para_below_heading,Point_1=Point_1,Point_2=Point_2,Point_3=Point_3,Point_4=Point_4)
        obj.save()

    return render(request,"pages/service_edit.html")


def Join_Our_Community_save(request):
    heading = request.POST.get("#Sub-heading")
    Completed_Projects = request.POST.get("#Completed_projects")
    Satisfied_customers = request.POST.get("#Satisfied_customers")
    Expert_Employees = request.POST.get("#Expert_Employees")

    obj = JoinOurCommunity(heading=heading,Completed_Projects=Completed_Projects,Satisfied_customers=Satisfied_customers,Expert_Employees=Expert_Employees)
    obj.save()

def demoday(request):
    return render(request,"demoday.html",{'DemoDayPic':get_DemoDayPic(),'DemoDayTOPSECTION':DemoDayTOPSECTION.objects.all()[::-1]})

def demoday_edit(request):
    return render(request,"pages/demoday.html",{'DemoDayPic':get_DemoDayPic(),'DemoDayTOPSECTION':DemoDayTOPSECTION.objects.all()[::-1]})

def DemoDayTOPSECTION_save(request):
    heading = request.POST.get("#heading")
    Content = request.POST.get("#Content")
    obj = DemoDayTOPSECTION(Content=Content,heading=heading)
    obj.save()
    return render(request,"pages/demoday.html")

def DemoDayPic_save(request):
    categories = request.POST.get("#heading1")
    image = request.FILES['#image_load']
    obj = DemoDayPic(image=image,categories=categories)
    obj.save()
    return render(request,"pages/demoday.html")

def delete_DemoDayPic(request):
    id = request.POST.get("id")
    size = len(DemoDayPic.objects.all())
    if size > 1:
        image = DemoDayPic.objects.get(id=id)
        image.delete()
    else:
        pass
    return render(request,"home/logo.html")

def set_DemoDayPic(request):
    id = request.POST.get("id")
    image = DemoDayPic.objects.get(id=id)
    image_ = image.image
    reson = image.categories
    image.delete() 
    obj = DemoDayPic(image=image_,categories=reson)
    obj.save()
    return render(request,"home/logo.html")


def stategovtfunds (request):
    return render(request,"stategovtfund.html")

def stategovtfunds_edit (request):
    return render(request,"pages/state_govt_fund_edit.html",{'StateGovtFund':StateGovtFund.objects.all()[::-1],'StateGovtFundSecondSection':StateGovtFundSecondSection.objects.all()[::-1],'StateGovtFundEligibilitySection':StateGovtFundEligibilitySection.objects.all()[::-1]})

def StateGovtFund_save(request):
    Main_Heading = request.POST.get("#Main_heading")
    Sub_Heading = request.POST.get("#Sub_Heading")
    Point_1 = request.POST.get("#Point_1")
    Point_2 = request.POST.get("#Point_2")
    Point_3 = request.POST.get("#Point_3")

    obj = StateGovtFund(Main_Heading=Main_Heading,Sub_Heading=Sub_Heading,Point_1=Point_1,Point_2=Point_2,Point_3=Point_3)
    obj.save()

    return render(request,"pages/state_govt_fund_edit.html")

def StateGovtFundSecondSection_save(request):
    Main_Heading = request.POST.get("#Main_heading")
    Sub_Heading = request.POST.get("#Sub-heading")
    Point_1 = request.POST.get("#Point-1")
    Point_2 = request.POST.get("#Point-2")
    Point_3 = request.POST.get("#Point-3")

    try:
        image = request.FILES['#fileInput-single11']
        obj = StateGovtFundSecondSection(image=image,Main_Heading=Main_Heading,Sub_Heading=Sub_Heading,Point_1=Point_1,Point_2=Point_2,Point_3=Point_3)
        obj.save()
    except:
        obj = StateGovtFundSecondSection.objects.all()[::-1][0]
        obj = StateGovtFundSecondSection(image=obj.image,Main_Heading=Main_Heading,Sub_Heading=Sub_Heading,Point_1=Point_1,Point_2=Point_2,Point_3=Point_3)
        obj.save()

    print("saved")

    return render(request,"pages/state_govt_fund_edit.html")


def StateGovtFundEligibilitySection_save(request):
    Sub_Heading = request.POST.get("#sub_heading")
    Point_1 = request.POST.get("#point_1")
    Point_2 = request.POST.get("#point_2")
    Point_3 = request.POST.get("#point_3")

    obj = StateGovtFundEligibilitySection(Sub_Heading=Sub_Heading,Point_1=Point_1,Point_2=Point_2,Point_3=Point_3)
    obj.save()

    print('saved')

    return render(request,"pages/state_govt_fund_edit.html")




def startuptn (request):
    return render(request,"startupTN.html")

def startuptn_edit (request):
    return render(request,"pages/startupTN_edit.html",{'Start_UpTN':Start_UpTN.objects.all()[::-1],'Start_UpTNContent2':Start_UpTNContent2.objects.all()[::-1],'Start_UpTNimg1':Start_UpTNimg1.objects.all()[::-1],'Start_UpTNimg2':Start_UpTNimg2.objects.all()[::-1]})

def Start_UpTN_save(request):
    Heading = request.POST.get("#heading")
    Content_1 = request.POST.get("#Content")
    obj = Start_UpTN(Heading=Heading,Content_1=Content_1)
    obj.save()
    return render(request,"pages/startupTN_edit.html")


def Start_UpTNContent2_save(request):
    Content_2 = request.POST.get("#heading1")
    obj = Start_UpTNContent2(Content_2=Content_2)
    obj.save()
    return render(request,"pages/startupTN_edit.html")

def Start_UpTNimg1_save(request):
    image = request.FILES['#fileInput-single11']
    obj = Start_UpTNimg1(image=image)
    obj.save()
    return render(request,"pages/startupTN_edit.html")

def Start_UpTNimg2_save(request):
    image = request.FILES['#fileInput-single2']
    obj = Start_UpTNimg2(image=image)
    obj.save()
    print("saved")
    return render(request,"pages/startupTN_edit.html")



def CentralGovernmentFunding (request):
    data = CentralGovernmentFundingDB.objects.all()[::-1]
    print(data)
    return render(request,"newventures.html",{'mentor':data,'sample':'hi'})

def CentralGovernmentFunding_edit(request):
    return render(request,"edtior/CentralGovernmentFunding_edit.html",{'mentor':CentralGovernmentFundingDB.objects.all()[::-1]})

def CentralGovernmentFunding_save(request):
    content = request.POST.get('#content')
    obj = CentralGovernmentFundingDB(Content=content)
    obj.save()
    for i in CentralGovernmentFundingDB.objects.all():
        print(i.Content)
    print("saved.....................................////////////////////////")
    return render(request,"edtior/CentralGovernmentFunding_edit.html")
